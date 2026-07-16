import csv
import os
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from models.compras import Factura
from models.emisor import Emisor
from models.Clasificaciones.tipo_documento import TipoDocumento
from models.Clasificaciones.clase_documento import ClaseDocumento
from models.emisor import Emisor
import tempfile

from datetime import datetime

def _obtener_datos(periodo_actual):

    if not periodo_actual:
        return []

    return Factura.obtener_por_periodo(periodo_actual)

def exportar_csv(periodo_actual):

    nombre = datetime.now().strftime("ANEXO_COMPRAS_%Y%m%d_%H%M%S.xlsx")
    datos = _obtener_datos(periodo_actual)

    if not datos:
        return False

    archivo = Path(tempfile.gettempdir()) / nombre

    wb = Workbook()
    ws = wb.active
    ws.title = "ANEXO DE COMPRAS"

    fila = 1

    for factura in datos:

        emisor = Emisor.obtener_datos_por_id(factura[17])

        nit = (emisor[4] or "").strip()
        nrc = (emisor[6] or "").strip()

        documento = nit if nit else nrc
    
        fecha = factura[2]

        if isinstance(fecha, str):
            try:
                fecha = datetime.strptime(fecha, "%Y-%m-%d")
            except:
                pass

        valores = [

            fecha,

            factura[18],
            factura[16],

            factura[3],

            documento,

            emisor[3],

            float(factura[10]),
            float(factura[11]),
            float(factura[12]),

            float(factura[6]),

            float(factura[13]),
            float(factura[14]),
            float(factura[15]),

            float(factura[7]),

            float(factura[8]),

            factura[22],
            factura[19],
            factura[20],
            factura[21],

            3,
        ]

        for col, valor in enumerate(valores, start=1):

            celda = ws.cell(fila, col)

            celda.value = valor

            if col == 1:
                celda.number_format = "d/m/yyyy"

            elif 7 <= col <= 15:
                celda.number_format = "0.00"

            elif col in (2, 3, 16, 17, 18, 19, 20):
                celda.number_format = "0"

            if col in (4, 5):  # Código de Generación y NIT/NRC
                celda.value = str(valor)
                celda.number_format = "@"

        fila += 1

    for i in range(1,21):
        ws.column_dimensions[get_column_letter(i)].width = 18

    wb.save(archivo)

    os.startfile(archivo)

    return archivo

def exportar_excel(periodo_actual):

    nombre = datetime.now().strftime("compras_%Y%m%d_%H%M%S.xlsx")
    datos = _obtener_datos(periodo_actual)

    if not datos:
        return False

    # ruta = Path("exports/compras")
    # ruta.mkdir(parents=True, exist_ok=True)

    # archivo = ruta / nombre
    archivo = Path(tempfile.gettempdir()) / nombre

    wb = Workbook()
    ws = wb.active
    ws.title = "Compras"

    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter

    # =====================================================
    # ESTILOS
    # =====================================================

    color_principal = "2F5597"
    color_encabezado = "EDEDED"
    color_borde = "D9D9D9"
    color_fila = "FAFAFA"

    borde = Border(
        left=Side(style="thin", color=color_borde),
        right=Side(style="thin", color=color_borde),
        top=Side(style="thin", color=color_borde),
        bottom=Side(style="thin", color=color_borde),
    )

    fuente_titulo = Font(
        name="Calibri",
        size=18,
        bold=True,
        color=color_principal
    )

    fuente_subtitulo = Font(
        name="Calibri",
        size=10,
        color="666666"
    )

    fuente_encabezado = Font(
        name="Calibri",
        size=11,
        bold=True,
        color="000000"
    )

    fuente_normal = Font(
        name="Calibri",
        size=11
    )

    fuente_total = Font(
        name="Calibri",
        size=11,
        bold=True
    )

    relleno_encabezado = PatternFill(
        "solid",
        fgColor=color_encabezado
    )

    relleno_alterno = PatternFill(
        "solid",
        fgColor=color_fila
    )

    # =====================================================
    # TITULO
    # =====================================================

    ws.merge_cells("A1:I1")

    c = ws["A1"]
    c.value = "REPORTE DE COMPRAS"
    c.font = fuente_titulo

    ws["A2"] = "Período"
    ws["B2"] = str(periodo_actual)

    ws["G2"] = "Generado"

    ws["H2"] = datetime.now().strftime("%d/%m/%Y %H:%M")

    ws["A2"].font = fuente_subtitulo
    ws["B2"].font = fuente_subtitulo
    ws["G2"].font = fuente_subtitulo
    ws["H2"].font = fuente_subtitulo

    # =====================================================
    # ENCABEZADOS
    # =====================================================

    encabezados = [
        "Fecha",
        "Clase Documento",
        "Tipo Documento",
        "Código Generación",

        "NIT O NRC del Proveedor",
        "Nombre del Proveedor",

        "Compras Internas Exentas",
        "Internaciones Exentas Y/O No Sujetas",
        "Importaciones Exentas Y/O No Sujetas",
        "Compras Gravadas Internas",
        "Internaciones Gravadas de Bienes",
        "Importaciones Gravadas de Bienes",
        "Importaciones Gravadas de Servicios",
        "Credito Fiscal",
        "Total de Compras",

        "Tipo de Operacion",
        "Clasificacion",
        "Sector",
        "Tipo de Costo/Gasto",
        "Numero de Anexo"
    ]

    fila_inicio = 4

    for col, texto in enumerate(encabezados, start=1):

        celda = ws.cell(fila_inicio, col)

        celda.value = texto
        celda.font = fuente_encabezado
        celda.fill = relleno_encabezado
        celda.alignment = Alignment(horizontal="center", vertical="center")
        celda.border = borde

    # =====================================================
    # DATOS
    # =====================================================

    fila = fila_inicio + 1

    for factura in datos:

        emisor = Emisor.obtener_datos_por_id(factura[17])
        nit = (emisor[4] or "").strip()
        nrc = (emisor[6] or "").strip()
        NUMERO_ANEXO = 3

        documento_proveedor = nit if nit else nrc

        valores = [

            str(factura[2]), #Fecha
            ClaseDocumento.obtener_por_id(factura[18]), #Clase de Documento
            TipoDocumento.obtener_por_id(factura[16]), #Tipo de Documento
            str(factura[3]), #Codigo de Generacion
            documento_proveedor, #NIT
            emisor[3], #Nombre del proveedor

            float(factura[10]), #Compras internas exentas
            float(factura[11]), #internaciones exentas no sujetas
            float(factura[12]), #importaciones exentas no sujetas

            float(factura[6]), #subtotal

            float(factura[13]), #internaciones gravadas bienes
            float(factura[14]), #importaciones gravadas bienes
            float(factura[15]), #importaciones gravadas servicios

            float(factura[7]), #IVA
            float(factura[8]), #total

            float(factura[22]), #Tipo Operacion
            float(factura[19]), #clasificacion
            float(factura[20]), #sector
            float(factura[21]), #Tipo costo/gasto
            NUMERO_ANEXO, #Numero de anexo
        ]

        for col, valor in enumerate(valores, start=1):

            celda = ws.cell(fila, col)

            celda.value = valor
            celda.font = fuente_normal
            celda.border = borde

            if fila % 2 == 0:
                celda.fill = relleno_alterno

            if col == 2:
                celda.alignment = Alignment(horizontal="center")

            elif  7 <= col <= 15:
                celda.alignment = Alignment(horizontal="right")
                celda.number_format = '#,##0.00'

            elif col in (16, 17, 18, 19, 20):
                celda.alignment = Alignment(horizontal="center")
                celda.number_format = '0'

            else:
                celda.alignment = Alignment(horizontal="left")

        fila += 1

    # =====================================================
    # TOTAL
    # =====================================================

    ws.cell(fila + 1, 9).value = "TOTAL"

    ws.cell(fila + 1, 9).font = fuente_total

    ws.cell(fila + 1, 9).alignment = Alignment(horizontal="right")

    ws.cell(fila + 1, 10).value = f"=SUM(J5:J{fila-1})"

    ws.cell(fila + 1, 10).font = fuente_total

    ws.cell(fila + 1, 10).number_format = '$ #,##0.00'

    ws.cell(fila + 1, 9).border = borde
    ws.cell(fila + 1, 10).border = borde

    # =====================================================
    # AUTOAJUSTAR COLUMNAS
    # =====================================================

    for columna in ws.columns:

        longitud = 0

        letra = get_column_letter(columna[0].column)

        for celda in columna:

            if celda.coordinate in ws.merged_cells:
                continue

            if celda.value:

                longitud = max(longitud, len(str(celda.value)))

        ws.column_dimensions[letra].width = min(longitud + 4, 45)

    # =====================================================
    # CONFIGURACIÓN
    # =====================================================

    ws.freeze_panes = "A5"

    ws.auto_filter.ref = f"A4:I{fila-1}"

    ws.sheet_view.showGridLines = False

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[4].height = 24

    wb.save(archivo)

    os.startfile(archivo)

    return archivo

def exportar_casilla_163(periodo_actual):

    nombre = datetime.now().strftime("CASILLA_163_%Y%m%d_%H%M%S.xlsx")
    datos = _obtener_datos(periodo_actual)

    if not datos:
        return False

    archivo = Path(tempfile.gettempdir()) / nombre

    wb = Workbook()
    ws = wb.active
    ws.title = "CASILLA 163"

    fila = 1

    for factura in datos:

        emisor = Emisor.obtener_datos_por_id(factura[17])

        nit = (emisor[4] or "").strip().replace("-", "")
        nrc = (emisor[6] or "").strip().replace("-", "")
        dui = (emisor[5] or "").strip()

        documento = nit if nit else nrc

        fecha = factura[2]

        if isinstance(fecha, str):
            try:
                fecha = datetime.strptime(fecha, "%Y-%m-%d")
            except:
                pass

        valores = [

            documento,          # A - NIT/NRC
            fecha,              # B - Fecha
            factura[16],        # C - Tipo Documento
            factura[5],         # D - Sello de Recepcion
            str(factura[3]).replace("-", ""),  # Código Generación sin guiones
            float(factura[6]),  # F - Subtotal
            float(factura[7]),  # G - IVA
            dui,                 # H - Vacía
            8                   # I - Casilla

        ]

        for col, valor in enumerate(valores, start=1):

            celda = ws.cell(fila, col)
            celda.value = valor

            # NIT / NRC
            if col == 1:
                celda.value = str(valor)
                celda.number_format = "@"

            # Fecha
            elif col == 2:
                celda.number_format = "d/m/yyyy"

            # Tipo de documento y Casilla
            elif col in (3, 9):
                celda.number_format = "0"

            # Número de control y Código de generación
            elif col in (4, 5):
                celda.value = str(valor)
                celda.number_format = "@"

            # Subtotal e IVA
            elif col in (6, 7):
                celda.number_format = "0.00"

        fila += 1

    # Todas las columnas con el mismo ancho
    for i in range(1, 10):
        ws.column_dimensions[get_column_letter(i)].width = 22

    wb.save(archivo)

    os.startfile(archivo)

    return archivo